import pandas as pd
import numpy as np
import glob
import openpyxl
from openpyxl.styles import Font
from openpyxl import load_workbook
from openpyxl import Workbook
from openpyxl.styles import Border, Side
wb = Workbook()


from datetime import datetime
start_time = datetime.now()


def set_border(cell_range,ws):
    thin = Side(border_style="medium",color="000000")
    for row in ws[cell_range]:
        for cell in row:
            cell.border = Border(top=thin, left=thin, right=thin, bottom=thin)
            
# function for highlighting cells
def set_highlight(r,c,ws):
    # print(r,c)
#     wb.active
    cell=ws.cell(row = r, column = c)
    my_high = openpyxl.styles.colors.Color(rgb='00FFFF00')
    my_fill = openpyxl.styles.fills.PatternFill(patternType='solid', fgColor=my_high)
    cell.fill = my_fill



def OctantCount(l, r, Octant):
    thisdict={'+1':0, '-1':0, '+2':0, '-2':0, '+3':0, '-3':0, '+4':0, '-4':0}
    for i in range(l, r+1):
#         print(i)
        thisdict[Octant[i]]+=1
    return thisdict

def GiveRank(thisdict):
    curdict={}
    l1=['Rank of +1', 'Rank of -1', 'Rank of +2', 'Rank of -2', 'Rank of +3', 'Rank of -3', 'Rank of +4', 'Rank of -4']
    l2=list(thisdict.values())
    pairlist=list(zip(l1, l2))
    pairlist.sort(key=lambda x: x[1])
    pairlist.reverse()
    i=0
    while i<8:
        curval=pairlist[i][1]
        curdict[pairlist[i][0]]=str(i+1)
        i+=1
    return curdict

def TransitionCount(l, r, Octant):
    thisdict={' +1': 0, ' -1': 0, ' +2': 0, ' -2': 0, ' +3': 0, ' -3': 0, ' +4': 0, ' -4': 0}
    finaldict={' +1':thisdict, ' -1':thisdict, ' +2':thisdict, ' -2':thisdict, ' +3':thisdict, ' -3':thisdict, ' +4':thisdict, ' -4':thisdict}
    #finaldict[i][j] means trans from i to j
    for i in range(l, r):
        _from=' '+Octant[i]
        _to=' '+Octant[i+1]
        finaldict[_from][_to]+=1
    if r+1<len(Octant):
        _from=' '+Octant[r]
        _to=' '+Octant[r+1]
        finaldict[_from][_to]+=1
    return finaldict


#Help
def octant_analysis(mod=5000):
	filenames = glob.glob(".\input\*.xlsx")
	for file in filenames:
		f_name=str(file)
		f_name=f_name.replace(".xlsx","")
		f_name=f_name[8:]
		j=len(f_name)-1
		s_name=""
		print("Reading file = ",file)
		str_mod=str(mod)
		stringg=f_name+"vel_octant_analysis_mod%s"%str_mod
		stringg1=stringg
		df=pd.read_excel(file)
		# df=pd.read_excel(r'C:\Users\Dell Vostro 3491\OneDrive\Documents\GitHub\2001ME42_2022\tut07\input\5.3.xlsx')
		# set max columns to none
		pd.set_option("display.max_columns", None)
		# set colwidth hidher
		pd.set_option('display.max_colwidth', 100)
		df = df.reindex(columns = df.columns.tolist() + ['U Avg', 'V Avg', 'W Avg', "U'=U - U avg", "V'=V - V avg", "W'=W - W avg", 'Octant', ' ', '  ', 'Overall Octant Count', '+1', '-1', '+2', '-2', '+3', '-3', '+4', '-4', 'Rank of +1', 'Rank of -1', 'Rank of +2', 'Rank of -2', 'Rank of +3', 'Rank of -3', 'Rank of +4', 'Rank of -4', 'Octant ID of Rank 1', 'Octant name of Rank 1', '   ', '    ', 'Overall Transition Count', ' +1', ' -1', ' +2', ' -2', ' +3', ' -3', ' +4', ' -4', '     ', 'Longest Subsquence Length', '      ', '       ', '        ', 'Longest Subsquence Length with Range', '         ', '          '])
		U_Avg=df['U'].mean()
		# print(U_Avg)
		U_Avg=round(U_Avg, 3)#setting precision to 8
		V_Avg=df['V'].mean()
		V_Avg=round(V_Avg, 3)
		W_Avg=df['W'].mean()
		W_Avg=round(W_Avg, 3)
		U_Avg=str(U_Avg)
		V_Avg=str(V_Avg)
		W_Avg=str(W_Avg)
		df.at[0, 'U Avg'] = U_Avg
		df.at[0, 'V Avg'] = V_Avg
		df.at[0, 'W Avg'] = W_Avg
		U_Avg=float(U_Avg)
		V_Avg=float(V_Avg)
		W_Avg=float(W_Avg)
		df["U'=U - U avg"]=df['U']-U_Avg
		df["V'=V - V avg"]=df['V']-V_Avg
		df["W'=W - W avg"]=df['W']-W_Avg
		df["U'=U - U avg"] = df["U'=U - U avg"].apply(lambda x: format(float(x),".3f"))
		df["V'=V - V avg"] = df["V'=V - V avg"].apply(lambda x: format(float(x),".3f"))
		df["W'=W - W avg"] = df["W'=W - W avg"].apply(lambda x: format(float(x),".3f"))
		U_list=df["U'=U - U avg"].tolist()
		V_list=df["V'=V - V avg"].tolist()
		W_list=df["W'=W - W avg"].tolist()
		n=len(U_list)
		for i in range(n):
		    U_list_i=float(U_list[i])
		    V_list_i=float(V_list[i])
		    W_list_i=float(W_list[i])
		    if U_list_i>=0.0 and V_list_i>=0.0:
		        if W_list_i>0:
		            df.at[i, 'Octant']="+1"
		        else:
		            df.at[i, 'Octant']="-1"
		    elif U_list_i<0.0 and V_list_i>=0.0:
		        if W_list_i>0:
		            df.at[i, 'Octant']="+2"
		        else:
		            df.at[i, 'Octant']="-2"
		    elif U_list_i<0.0 and V_list_i<0.0:
		        if W_list_i>0:
		            df.at[i, 'Octant']="+3"
		        else:
		            df.at[i, 'Octant']="-3"
		    elif U_list_i>=0.0 and V_list_i<0.0:
		        if W_list_i>0:
		            df.at[i, 'Octant']="+4"
		        else:
		            df.at[i, 'Octant']="-4"
		cur_index=1


		#Overall Octant Count and Rank
		df.at[cur_index, 'Overall Octant Count']='Octant ID'
		df.at[cur_index, '+1']=df.at[cur_index, ' +1']='+1'
		df.at[cur_index, '-1']=df.at[cur_index, ' -1']='-1'
		df.at[cur_index, '+2']=df.at[cur_index, ' +2']='+2'
		df.at[cur_index, '-2']=df.at[cur_index, ' -2']='-2'
		df.at[cur_index, '+3']=df.at[cur_index, ' +3']='+3'
		df.at[cur_index, '-3']=df.at[cur_index, ' -3']='-3'
		df.at[cur_index, '+4']=df.at[cur_index, ' +4']='+4'
		df.at[cur_index, '-4']=df.at[cur_index, ' -4']='-4'
		st_high=[]
		Oct_list=df['Octant'].tolist()
		thisdict=OctantCount(0, n-1, Oct_list)
		rankdict=GiveRank(thisdict)
		cur_index+=1
		df.at[cur_index, 'Overall Octant Count']='Overall Count'
		OctantName={
		        '+1':'Internal outward interaction',
		        '-1':'External outward interaction',
		        '+2':'External Ejection',
		        '-2':'Internal Ejection',
		        '+3':'External inward interaction',
		        '-3':'Internal inward interaction',
		        '+4':'Internal sweep',
		        '-4':'External sweep'
		    }
		countdict={
		        '+1':0,
		        '-1':0,
		        '+2':0,
		        '-2':0,
		        '+3':0,
		        '-3':0,
		        '+4':0,
		        '-4':0
		    }
		rank_1=''
		cur_col=23
		cnt=0
		for i in thisdict:
			cnt+=1
			df.at[cur_index, i]=thisdict[i]
			txt='Rank of {}'
			txt=txt.format(i)
			df.at[cur_index, txt]=rankdict[txt]
			if rankdict[txt]=='1':
				cur_col+=cnt
				df.at[cur_index, 'Octant ID of Rank 1']=i
				df.at[cur_index, 'Octant name of Rank 1']=OctantName[i]
		st_high.append([cur_index, cur_col-1])
		cur_index+=1
		# print(thisdict)
		x=5000
		l=[]
		cur=0
		row_cnt=len(df.index)
		while cur<=row_cnt:
		    l.append(cur)
		    cur+=x
		l_n=len(l)
		for i in range(l_n):
		    low=l[i]
		    high=low+x-1
		    if high>=n:
		        high=n-1
		    thisdict=OctantCount(low, high, Oct_list)
		    rankdict=GiveRank(thisdict)
		    txt="{}-{}"
		    txt=txt.format(low, high)
		    df.at[cur_index, 'Overall Octant Count']=txt
		    cur_col=23
		    cnt=0
		    for i in thisdict:
		        cnt+=1
		        df.at[cur_index, i]=thisdict[i]
		        txt='Rank of {}'
		        txt=txt.format(i)
		        df.at[cur_index, txt]=rankdict[txt]
		        if rankdict[txt]=='1':
		#             print(txt, i)
		            df.at[cur_index, 'Octant ID of Rank 1']=i
		            df.at[cur_index, 'Octant name of Rank 1']=OctantName[i]
		            countdict[i]+=1
		            cur_col+=cnt
		    st_high.append([cur_index, cur_col-1])
		    cur_index+=1
		cur_index=10
		df.at[cur_index, 'Rank of +4']='Octant ID'
		df.at[cur_index, 'Rank of -4']='Octant Name'
		df.at[cur_index, 'Octant ID of Rank 1']='Count of Rank 1 Mod Values'
		cur_index+=1
		for i in countdict:
		    df.at[cur_index, 'Rank of +4']=i
		    df.at[cur_index, 'Rank of -4']=OctantName[i]
		    df.at[cur_index, 'Octant ID of Rank 1']=countdict[i]
		    cur_index+=1
		#     print(low, high)
		#     print(thisdict)
		



		#Transitions
		#'    ', 'Overall Transition Count', ' +1'
		transitiondict=TransitionCount(0, n-1, Oct_list)
		df.at[1, 'Overall Transition Count']='Octant #'
		df.at[0, ' +1']='To'
		df.at[2, '    ']='From'
		# print(transitiondict)
		cur_index=2
		for i in transitiondict:
		    df.at[cur_index, 'Overall Transition Count']=i[1:];
		    for j in transitiondict[i]:
		        df.at[cur_index, j]=transitiondict[i][j]
		    cur_index+=1
		cur_index+=1
		for k in range(l_n):
		    cur_index+=2
		    df.at[cur_index, 'Overall Transition Count']='Mod Transition Count'
		    cur_index+=1
		    low=l[k]
		    high=low+x-1
		    if high>=n:
		        high=n-1
		    txt="{}-{}"
		    txt=txt.format(low, high)
		    df.at[cur_index, 'Overall Transition Count']=txt
		    cur_index+=1
		    df.at[cur_index, 'Overall Transition Count']='Octant #'
		    df.at[cur_index-1, ' +1']='To'
		    df.at[cur_index+1, '    ']='From'
		    transitiondict=TransitionCount(low, high, Oct_list)
		    for i in transitiondict:
		        df.at[cur_index, i]=i[1:];
		    cur_index+=1
		    for i in transitiondict:
		        df.at[cur_index, 'Overall Transition Count']=i[1:];
		        for j in transitiondict[i]:
		            df.at[cur_index, j]=transitiondict[i][j]
		        cur_index+=1

		        
		        
		        

		#Longest Subsequence Length        
		#'Longest Subsquence Length', '      ', '       ', '        '
		#'Longest Subsquence Length with Range', '         ', '          '        
		    #creating a dictionary having key each octant ID and value list of size 2
		    #having first value as Longest Subsequence Length and second value as 
		    #frequency of Longest Subsequence Length
		thisdict={
		    '+1':[0, 0],
		    '-1':[0, 0],
		    '+2':[0, 0],
		    '-2':[0, 0],
		    '+3':[0, 0],
		    '-3':[0, 0],
		    '+4':[0, 0],
		    '-4':[0, 0]
		}
		    #creating a dictionary having key each octant ID and value a list storing 
		    #all the starting indices of Longest Subsequence Length
		newdict={
		    '+1':[],
		    '-1':[],
		    '+2':[],
		    '-2':[],
		    '+3':[],
		    '-3':[],
		    '+4':[],
		    '-4':[]
		}
		i=0
		while i<n:
		    cur_val=Oct_list[i]
		    cur_freq=1
		    j=i+1
		    while j<n :
		        if Oct_list[j]!=cur_val:
		            break
		        cur_freq+=1
		        j+=1
		    if thisdict[cur_val][0]==cur_freq:
		        thisdict[cur_val][1]+=1
		        newdict[cur_val].append(i)
		    elif thisdict[cur_val][0]<cur_freq:
		        newdict[cur_val].clear()
		        newdict[cur_val].append(i)
		        thisdict[cur_val][1]=1
		        thisdict[cur_val][0]=cur_freq
		    i=j


		#'Longest Subsquence Length with Range', '         ', '          '  
		cur_index=1
		df.at[cur_index, 'Longest Subsquence Length']='Octant##'
		df.at[cur_index, '      ']='Longest Subsquence Length'
		df.at[cur_index, '       ']='Count'
		df.at[cur_index, 'Longest Subsquence Length with Range']='Octant###'
		df.at[cur_index, '         ']='Longest Subsquence Length'
		df.at[cur_index, '          ']='Count'
		cur_index+=1
		for i in thisdict:
		    df.at[cur_index, 'Longest Subsquence Length']=i
		    df.at[cur_index, '      ']=thisdict[i][0]
		    df.at[cur_index, '       ']=thisdict[i][1]
		    cur_index+=1

		cur_index=2
		for i in newdict:
		    #df['Count_2'][cur_index]=i
		    df.at[cur_index, 'Longest Subsquence Length with Range']=i
		#     df['Longest Subsequence Length_2'][cur_index]=thisdict[i][0]
		    df.at[cur_index, '         ']=thisdict[i][0]
		#     df['Count_3'][cur_index]=thisdict[i][1]
		    df.at[cur_index, '          ']=thisdict[i][1]
		    cur_index+=1
		    df.at[cur_index, 'Longest Subsquence Length with Range']='Time'
		    df.at[cur_index, '         ']='From'
		    df.at[cur_index, '          ']='To'
		    cur_index+=1
		    for j in newdict[i]:
		        df.at[cur_index, '         ']=df['T'][j]
		        df.at[cur_index, '          ']=df['T'][j+thisdict[i][0]-1]
		        cur_index+=1
		# df['id'] = df['id'].astype("string")     
		df.style.applymap(lambda x: "background-color: red" if x==1 else "background-color: white")
		# df.head()
		df.to_excel('./output/%s.xlsx'%stringg,index=False) #creating output excel file
		writer = load_workbook(filename='./output/%s.xlsx'%stringg1)
		worksheet = writer['Sheet1']
		for j in range(len(st_high)):
		    set_highlight(st_high[j][0]+2,st_high[j][1],worksheet)

		#         workbook = xlsxwriter.Workbook('%s.xlsx'%stringg1)

		set_border('N3:AF7',worksheet)
		set_border('AI3:AQ11',worksheet)
		set_border('AC9:AE17',worksheet)
		set_border('AI17:AQ25',worksheet)
		set_border('AS3:AU11',worksheet)
		set_border('N3:AF7',worksheet)
		set_border('AI31:AQ39',worksheet)
		set_border('AI46:AQ53',worksheet)
		set_border('AI59:AQ67',worksheet)
		set_border('AW3:AY28',worksheet)
		writer.save('./output/%s.xlsx'%stringg1)
##Read all the excel files in a batch format from the input/ folder. Only xlsx to be allowed
##Save all the excel files in a the output/ folder. Only xlsx to be allowed
## output filename = input_filename[_octant_analysis_mod_5000].xlsx , ie, append _octant_analysis_mod_5000 to the original filename. 

###Code

from platform import python_version
ver = python_version()

if ver == "3.8.10":
	print("Correct Version Installed")
else:
	print("Please install 3.8.10. Instruction are present in the GitHub Repo/Webmail. Url: https://pastebin.com/nvibxmjw")


mod=5000
octant_analysis(mod)






#This shall be the last lines of the code.
end_time = datetime.now()
print('Duration of Program Execution: {}'.format(end_time - start_time))
